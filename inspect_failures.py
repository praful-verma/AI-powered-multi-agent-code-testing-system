import openpyxl
from pathlib import Path

path = Path('runs/test_cases_verify.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for row in ws.iter_rows(min_row=2, values_only=True):
    data = dict(zip(headers, row))
    if data.get('status') in {'Fail', 'Error'}:
        print('---')
        for key in ['test_id', 'target_file', 'target_function_or_route', 'status', 'error_message', 'test_description']:
            print(f'{key}: {data.get(key)}')
        print()
