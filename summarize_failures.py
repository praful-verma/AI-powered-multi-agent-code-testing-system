import openpyxl
from pathlib import Path

p = Path('runs/test_cases_verify.xlsx')
wb = openpyxl.load_workbook(p)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data = dict(zip(headers, row))
    if data.get('status') in {'Fail', 'Error'}:
        rows.append(data)

print(len(rows))
for row in rows[:20]:
    print(f"{row.get('test_id')} {row.get('target_file')} {row.get('target_function_or_route')}")
