from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from multiagent_testing.schema import COLUMNS, SHEET_NAME


class ExcelStore:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def ensure_workbook(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = self._sheet(wb)
            changed = self._ensure_headers(ws)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append(COLUMNS)
            ws.freeze_panes = "A2"
            changed = True
        if changed:
            self._save(wb)

    def append_rows(self, rows: list[dict[str, Any]]) -> None:
        self.ensure_workbook()
        wb = load_workbook(self.path)
        ws = self._sheet(wb)
        self._ensure_headers(ws)
        for row in rows:
            ws.append([self._clean_cell_value(row.get(column, "")) for column in COLUMNS])
        self._save(wb)

    def upsert_rows(self, rows: list[dict[str, Any]], match_columns: list[str]) -> None:
        if not rows:
            return
        self.ensure_workbook()
        wb = load_workbook(self.path)
        ws = self._sheet(wb)
        self._ensure_headers(ws)
        header_map = {name: idx + 1 for idx, name in enumerate(self._headers(ws))}
        existing: dict[tuple[str, ...], int] = {}
        for row_idx in range(2, ws.max_row + 1):
            key = tuple(str(ws.cell(row=row_idx, column=header_map[column]).value or "") for column in match_columns)
            if any(key):
                existing[key] = row_idx

        for row in rows:
            key = tuple(str(row.get(column) or "") for column in match_columns)
            row_idx = existing.get(key)
            if row_idx is None or not any(key):
                ws.append([self._clean_cell_value(row.get(column, "")) for column in COLUMNS])
                continue
            existing_test_id = ws.cell(row=row_idx, column=header_map["test_id"]).value
            for column in COLUMNS:
                value = existing_test_id if column == "test_id" and existing_test_id else row.get(column, "")
                ws.cell(row=row_idx, column=header_map[column]).value = self._clean_cell_value(value)
        self._save(wb)

    def rows(self) -> list[dict[str, Any]]:
        self.ensure_workbook()
        wb = load_workbook(self.path)
        ws = self._sheet(wb)
        headers = self._headers(ws)
        data = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in excel_row):
                continue
            data.append({header: excel_row[idx] if idx < len(excel_row) else None for idx, header in enumerate(headers)})
        return data

    def update_by_test_id(self, test_id: str, updates: dict[str, Any]) -> None:
        self.bulk_update_by_test_id({test_id: updates})

    def bulk_update_by_test_id(self, updates: dict[str, dict[str, Any]]) -> None:
        if not updates:
            return
        self.ensure_workbook()
        wb = load_workbook(self.path)
        ws = self._sheet(wb)
        header_map = {name: idx + 1 for idx, name in enumerate(self._headers(ws))}
        test_id_col = header_map["test_id"]
        for row_idx in range(2, ws.max_row + 1):
            test_id = ws.cell(row=row_idx, column=test_id_col).value
            if test_id not in updates:
                continue
            for column, value in updates[test_id].items():
                if column not in header_map:
                    raise ValueError(f"Unknown Excel column: {column}")
                ws.cell(row=row_idx, column=header_map[column]).value = self._clean_cell_value(value)
        self._save(wb)

    def next_test_number(self) -> int:
        rows = self.rows()
        max_seen = 0
        for row in rows:
            test_id = str(row.get("test_id") or "")
            if test_id.startswith("TC-"):
                try:
                    max_seen = max(max_seen, int(test_id[3:]))
                except ValueError:
                    pass
        return max_seen + 1

    def _sheet(self, wb) -> Worksheet:
        return wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    def _headers(self, ws: Worksheet) -> list[str]:
        return [cell.value for cell in ws[1]]

    def _ensure_headers(self, ws: Worksheet) -> bool:
        headers = self._headers(ws)
        changed = False
        for column in COLUMNS:
            if column not in headers:
                ws.cell(row=1, column=len(headers) + 1).value = column
                headers.append(column)
                changed = True
        return changed

    def _save(self, wb) -> None:
        try:
            wb.save(self.path)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot write Excel workbook '{self.path}'. Close it in Excel/OneDrive preview, "
                "or rerun with --excel-path pointing to a new .xlsx file."
            ) from exc

    def _clean_cell_value(self, value: Any) -> Any:
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value
