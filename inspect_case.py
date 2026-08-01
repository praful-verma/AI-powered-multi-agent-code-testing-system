import openpyxl
from pathlib import Path

path = Path('runs/test_cases_verify.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for target in ['TC-0005','TC-0010','TC-0032','TC-0041','TC-0080','TC-0117']:
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if data.get('test_id') == target:
            print('===', target, '===')
            print(data.get('test_file_path'))
            print(data.get('test_code'))
            print()
            break
